import re
import pdfplumber
import unicodedata
from openpyxl import Workbook, load_workbook

# -----------------------------------------------------------
# Normalização de nomes (caso origem seja Excel)
# -----------------------------------------------------------

def norm(s):
    if s is None:
        return ""
    t = unicodedata.normalize("NFKD", str(s))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return " ".join(t.lower().strip().split())

# -----------------------------------------------------------
# Conversão de valores monetários
# -----------------------------------------------------------

def parse_money(v):
    if not v:
        return 0.0
    x = v.replace("R$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(x)
    except:
        return 0.0

# -----------------------------------------------------------
# Regex compatível com o formato REAL das linhas do PDF alvo
# -----------------------------------------------------------

REGEX_LINHA = re.compile(
    r"(?P<cpf_mask>\d{3}\.\d{3}\.\d{3}-\d{2})\s+"
    r"(?P<cpf_num>\d{11})\s+"
    r"(?P<nome>[A-ZÀ-ÖØ-öø-ÿ\s]+?)\s+"
    r"[A-ZÀ-ÖØ-öø-ÿ]+\s+"
    r"(?P<valor>R?\$?\s*\d{1,3}(?:\.\d{3})*,\d{2})"
)

# -----------------------------------------------------------
# Ler PDF (retorna lista de linhas)
# -----------------------------------------------------------

def ler_pdf(path):
    linhas = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for ln in txt.splitlines():
                linhas.append(" ".join(ln.split()))
    return linhas

# -----------------------------------------------------------
# Busca por CPF no PDF alvo
# -----------------------------------------------------------

def buscar_no_pdf_alvo(linhas_pdf, cpf_procurado):
    cpf_num = cpf_procurado.replace(".", "").replace("-", "")
    
    for linha in linhas_pdf:
        if cpf_num in linha.replace(".", "").replace("-", ""):
            m = REGEX_LINHA.search(linha)
            if m:
                return {
                    "cpf": m.group("cpf_mask"),
                    "matricula": m.group("cpf_num"),
                    "nome": m.group("nome").strip(),
                    "valor": m.group("valor")
                }
    return None

# -----------------------------------------------------------
# Origem em PDF (cpf + nome)
# -----------------------------------------------------------

def extrair_origem_pdf(pdf_path):
    linhas = ler_pdf(pdf_path)
    origem = []
    cpf_re = re.compile(r"\d{3}\.\d{3}\.\d{3}-\d{2}")

    for ln in linhas:
        m = cpf_re.search(ln)
        if m:
            cpf = m.group()
            nome = ln.replace(cpf, "").strip()
            origem.append({"cpf": cpf, "nome": nome})

    return origem

# -----------------------------------------------------------
# Origem em Excel
# -----------------------------------------------------------

def extrair_origem_excel(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    origem = []
    for r in range(2, ws.max_row + 1):
        cpf = ws.cell(row=r, column=1).value
        nome = ws.cell(row=r, column=2).value
        if cpf:
            origem.append({"cpf": str(cpf), "nome": nome})

    return origem

# -----------------------------------------------------------
# Criar planilha final
# -----------------------------------------------------------

def salvar_planilha_final(registros, saida_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["CPF", "Matrícula", "Nome", "Valor"])

    for rec in registros:
        ws.append([
            rec["cpf"],
            rec["matricula"],
            rec["nome"],
            parse_money(rec["valor"])
        ])

    wb.save(saida_path)
    print(f"\n✔ Planilha criada: {saida_path}")
    print(f"✔ Total de registros: {len(registros)}")

# -----------------------------------------------------------
# Menu principal
# -----------------------------------------------------------

def main():
    print("=== SCRIPT DE EXTRAÇÃO DE BENEFÍCIOS ===\n")
    print("Escolha a origem dos CPFs:")
    print("1 - PDF")
    print("2 - Excel (.xlsx)")
    opc = input("\nOpção: ").strip()

    if opc == "1":
        pdf_origem = input("PDF origem (CPF + nome): ").strip()
        origem = extrair_origem_pdf(pdf_origem)

    elif opc == "2":
        xlsx_origem = input("Planilha origem (.xlsx): ").strip()
        origem = extrair_origem_excel(xlsx_origem)

    else:
        print("Opção inválida.")
        return

    pdf_alvo = input("PDF alvo (onde estão os benefícios): ").strip()
    print("\nLendo PDF alvo...")
    linhas_pdf = ler_pdf(pdf_alvo)

    encontrados = []
    nao_encontrados = []

    print("\nIniciando busca...\n")

    for o in origem:
        res = buscar_no_pdf_alvo(linhas_pdf, o["cpf"])
        if res:
            encontrados.append(res)
            print(f"✓ Encontrado: {res['cpf']} - {res['nome']}")
        else:
            nao_encontrados.append(o)
            print(f"✗ Não encontrado: {o['cpf']} - {o['nome']}")

    print("\n----------------------------------------")
    print("NOMES NÃO ENCONTRADOS NO PDF ALVO:")
    print("----------------------------------------")

    if nao_encontrados:
        for item in nao_encontrados:
            print(f"- {item['cpf']}  {item['nome']}")
    else:
        print("Todos os nomes foram encontrados.")

    saida = input("\nNome do arquivo de saída (.xlsx): ").strip()
    if not saida.lower().endswith(".xlsx"):
        saida += ".xlsx"

    salvar_planilha_final(encontrados, saida)

# -----------------------------------------------------------

if __name__ == "__main__":
    main()
