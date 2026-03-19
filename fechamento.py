import os
import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict, Counter
from PyPDF2 import PdfReader, PdfWriter, PdfMerger
import csv

# ==========================
# 1️⃣ CONFIGURAÇÃO
# ==========================
base_dir = os.path.dirname(os.path.abspath(__file__))
pasta_comprovantes = os.path.join(base_dir, "Comprovantes")
pasta_contracheques = os.path.join(base_dir, "Contracheques")
pasta_saida = os.path.join(base_dir, "Fechamento_Final")
os.makedirs(pasta_saida, exist_ok=True)

FUZZY_CUTOFF = 0.65  # tolerância do match (0.65 = mais permissivo)

# ==========================
# 2️⃣ FUNÇÕES AUXILIARES
# ==========================
def normalizar(texto):
    """Remove acentos, pontuação e deixa maiúsculo."""
    texto = ''.join(
        c for c in unicodedata.normalize('NFD', texto.upper())
        if unicodedata.category(c) != 'Mn'
    )
    texto = re.sub(r'[^A-Z0-9\s]', ' ', texto)
    texto = re.sub(r'\s+', ' ', texto)
    return texto.strip()

def similaridade(a, b):
    """Similaridade por sequência (ordem importa)."""
    return SequenceMatcher(None, a, b).ratio()

def calcular_score_nomes(nome1, nome2):
    """
    Calcula score considerando abreviações e ignorando partículas.
    Ex: 'IRISMAR F REIS' == 'IRISMAR FRANCISCO DOS REIS'
    """
    # 1. Score base (fuzzy string matching)
    score_seq = SequenceMatcher(None, nome1, nome2).ratio()
    if score_seq > 0.95:
        return score_seq

    # 2. Tokenização inteligente
    ignorados = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E'}
    
    t1 = [t for t in nome1.split() if t not in ignorados]
    t2 = [t for t in nome2.split() if t not in ignorados]
    
    if not t1 or not t2:
        return score_seq

    matches = 0
    used_t2 = set()
    
    # Tenta casar cada token de t1 com um token de t2
    for token1 in t1:
        for i, token2 in enumerate(t2):
            if i in used_t2:
                continue
            
            # A) Match Exato
            if token1 == token2:
                matches += 1
                used_t2.add(i)
                break
            
            # B) Abreviação (ex: F vs FRANCISCO)
            # Verifica se um é inicial do outro
            if (len(token1) == 1 and token2.startswith(token1)) or \
               (len(token2) == 1 and token1.startswith(token2)):
                matches += 1
                used_t2.add(i)
                break
    
    # Score = proporção de tokens casados sobre o maior número de tokens
    # Ex: "JOAO S" (2) vs "JOAO SILVA" (2) -> 2 matches / 2 = 1.0
    score_tokens = matches / max(len(t1), len(t2))
    
    return max(score_seq, score_tokens)

def extrair_info_pagador(texto):
    """
    Extrai e normaliza informações do pagador.
    Mantém apenas os dois primeiros nomes.
    """
    m = re.search(r"INFORMA[ÇC][ÕO]ES\s+FORNECIDAS\s+PELO\s+PAGADOR\s*[:\.]?\s*([^\n]+)", texto, re.IGNORECASE)
    if m and m.group(1).strip():
        content = normalizar(m.group(1))
        tokens = content.split()
        if len(tokens) >= 2:
            return f"{tokens[0]} {tokens[1]}"
        elif len(tokens) == 1:
            return tokens[0]
    # Se o valor estiver na linha seguinte
    m2 = re.search(r"INFORMA[ÇC][ÕO]ES\s+FORNECIDAS\s+PELO\s+PAGADOR\s*[:\.]?\s*\n\s*([^\n]+)", texto, re.IGNORECASE)
    if m2:
        content = normalizar(m2.group(1))
        tokens = content.split()
        if len(tokens) >= 2:
            return f"{tokens[0]} {tokens[1]}"
        elif len(tokens) == 1:
            return tokens[0]
    return None

# NOVO: verificação estrita do nome
# - Exige compatibilidade do PRIMEIRO nome (obrigatória)
# - Permite abreviação apenas quando o token possui 1 letra
# - Ignora partículas como DE/DA/DO/DOS/DAS/E
def nome_compativel_estrito(a, b):
    ig = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E'}
    t1 = [t for t in a.split() if t not in ig]
    t2 = [t for t in b.split() if t not in ig]
    if len(t1) < 2 or len(t2) < 2:
        return False
    def tok_ok(x, y):
        if x == y:
            return True
        if len(x) == 1 and len(y) > 1 and y.startswith(x):
            return True
        if len(y) == 1 and len(x) > 1 and x.startswith(y):
            return True
        return False
    if not tok_ok(t1[0], t2[0]):
        return False
    if not tok_ok(t1[1], t2[1]):
        return False
    return True

def chave_nome2(nome):
    ig = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E'}
    t = [x for x in nome.split() if x not in ig]
    if len(t) < 2:
        return None
    return f"{t[0]} {t[1]}"

def eh_decimo_terceiro(pagador):
    p = normalizar(pagador or "")
    if not p:
        return False
    if re.search(r"\b(13|13O|13º|DECIMO|DECIMO TERCEIRO)\b", p):
        return True
    if p.startswith("13"):
        return True
    return False

def normalizar_contrato_chave(s):
    s = normalizar(s or "")
    if not s:
        return ""
    t = s.split()
    key = t[0] if t else ""
    if key.startswith("HEMO") or key == "HEMOCENTRO":
        return "HEMO"
    if key.startswith("SAUDE"):
        return "SAUDE"
    if key.startswith("HGG"):
        return "HGG"
    if key.startswith("SECULT"):
        return "SECULT"
    return key

def contrato_compativel(pagador, grupo):
    pk = normalizar_contrato_chave(pagador)
    gk = normalizar_contrato_chave(grupo)
    if not pk:
        return True
    contratos = {"HEMO", "HEMOCENTRO", "SAUDE", "HGG", "SECULT"}
    if gk not in contratos:
        return True
    if gk == "HEMOCENTRO":
        gk = "HEMO"
    return pk == gk

def pagador_compativel_arquivo(pagador, base_nome):
    base_norm = normalizar(base_nome or "")
    pag = normalizar(pagador or "")
    if not pag:
        return True
    toks = pag.split()
    chave = " ".join(toks[:2]) if toks else ""
    return bool(chave) and chave in base_norm

def canonicalizar_numero(s):
    n = re.sub(r'\D', '', s or '')
    n = n.lstrip('0')
    return n or '0'

def extrair_nome_empresa(pasta_contra, pasta_comp):
    padroes = [
        r"EMPRESA\s*:\s*([A-ZÀ-ÿ0-9\s\-\.\&]+)",
        r"EMPREGADOR\s*:\s*([A-ZÀ-ÿ0-9\s\-\.\&]+)",
        r"RAZ[ÃA]O\s+SOCIAL\s*:\s*([A-ZÀ-ÿ0-9\s\-\.\&]+)"
    ]
    def buscar(caminho):
        try:
            leitor = PdfReader(caminho, strict=False)
            for pagina in leitor.pages:
                texto = pagina.extract_text() or ""
                for pat in padroes:
                    m = re.search(pat, texto, re.IGNORECASE)
                    if m:
                        return normalizar(m.group(1))
        except Exception:
            pass
        return None
    for arquivo in os.listdir(pasta_contra):
        if arquivo.lower().endswith('.pdf'):
            nome = buscar(os.path.join(pasta_contra, arquivo))
            if nome:
                return nome
    for arquivo in os.listdir(pasta_comp):
        if arquivo.lower().endswith('.pdf'):
            nome = buscar(os.path.join(pasta_comp, arquivo))
            if nome:
                return nome
    return "EMPRESA"

def extrair_empresa_de_nome(base, cidade):
    m = re.search(r"(?i)recibo\s+de\s+pagamento(?:s)?\s*-\s*([^-]+?)\s*-\s*([^-]+)", base)
    empresa_raw, cidade_raw = None, None
    if m:
        empresa_raw = m.group(1).strip()
        cidade_raw = m.group(2).strip()
    else:
        partes = [p.strip() for p in re.split(r"\s*-\s*", base)]
        if len(partes) >= 3:
            empresa_raw = partes[-2]
            cidade_raw = partes[-1]
        elif len(partes) == 2:
            empresa_raw = partes[0]
            cidade_raw = partes[1]
        else:
            empresa_raw = partes[0]
            cidade_raw = ""
    empresa_norm = normalizar(empresa_raw) if empresa_raw else ""
    cidade_norm = normalizar(cidade_raw) if cidade_raw else ""
    chave_cidade = re.sub(r"\s+", "", normalizar(cidade))
    cidade_compact = re.sub(r"\s+", "", cidade_norm)
    if not cidade_norm or cidade_compact == chave_cidade:
        return empresa_norm or "EMPRESA"
    return empresa_norm or "EMPRESA"

def extrair_nomes_contracheques(pasta):
    cidades = defaultdict(list)
    for root, dirs, files in os.walk(pasta):
        for arquivo in files:
            if not arquivo.lower().endswith('.pdf'):
                continue
            
            rel_path = os.path.relpath(root, pasta)
            cidade_real = re.sub(r'(?i)[^A-Z]', '', os.path.splitext(arquivo)[0].split('-')[-1]).upper()
            
            if rel_path == ".":
                # Arquivo na raiz: grupo é a própria cidade (comportamento antigo)
                grupo = cidade_real
            else:
                # Arquivo em subpasta: grupo é o nome da pasta
                grupo = rel_path.upper()

            caminho = os.path.join(root, arquivo)
            try:
                leitor = PdfReader(caminho, strict=False)
                for i, pagina in enumerate(leitor.pages):
                    texto = pagina.extract_text() or ""
                    padroes_nome = [
                        r"FUNC\.?\s*:\s*\d+\s*-\s*([A-ZÀ-ÿ0-9\s\-\.]+)",
                        r"FUNCION[ÁA]RIO\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)",
                        r"NOME\s+DO\s+EMPREGADO\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)",
                        r"TRABALHADOR\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)"
                    ]
                    m = None
                    for pat in padroes_nome:
                        m = re.search(pat, texto, re.IGNORECASE)
                        if m:
                            break
                    mag = re.search(r"AG[ÊE]NCIA\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
                    mc = re.search(r"CONTA(?:\s+CORRENTE)?\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
                    ag = canonicalizar_numero(mag.group(1)) if mag else None
                    conta = canonicalizar_numero(mc.group(1)) if mc else None
                    if m:
                        nome = normalizar(m.group(1))
                        cidades[grupo].append((nome, caminho, i, ag, conta, cidade_real))
            except Exception as e:
                print(f"⚠️ Erro lendo {arquivo}: {e}")
    return cidades

def extrair_nomes_comprovantes(pasta):
    nomes = []
    arquivos = [f for f in os.listdir(pasta) if f.lower().endswith('.pdf')]
    if not arquivos:
        raise FileNotFoundError("Nenhum comprovante encontrado.")
    caminho = os.path.join(pasta, arquivos[0])
    try:
        leitor = PdfReader(caminho, strict=False)
        for i, pagina in enumerate(leitor.pages):
            texto = pagina.extract_text() or ""
            m = re.search(r"NOME\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)", texto, re.IGNORECASE)
            mag = re.search(r"AG[ÊE]NCIA\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
            mc = re.search(r"CONTA(?:\s+CORRENTE)?\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
            pagador = extrair_info_pagador(texto)

            ag = canonicalizar_numero(mag.group(1)) if mag else None
            conta = canonicalizar_numero(mc.group(1)) if mc else None
            if m:
                nome = normalizar(m.group(1))
                nomes.append((nome, caminho, i, ag, conta, pagador))
    except Exception as e:
        print(f"⚠️ Erro lendo comprovante: {e}")
    return nomes

# NOVO: exportação de faltantes para Excel
def exportar_faltantes_planilha(faltantes, caminho_xlsx):
    # Cria uma planilha estilizada; fallback para CSV caso openpyxl não esteja disponível
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "Faltantes"
        ws.append(["Nome", "Empresa", "Cidade"])
        for nome, empresa, cidade in sorted(faltantes):
            ws.append([nome, empresa, cidade])
        for c in range(1, 4):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
            cell.alignment = Alignment(horizontal="center")
        ws.auto_filter.ref = f"A1:C{ws.max_row}"
        ws.freeze_panes = "A2"
        thin = Side(style="thin", color="999999")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx in range(1, 4):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)
        wb.save(caminho_xlsx)
    except Exception:
        caminho_csv = os.path.splitext(caminho_xlsx)[0] + ".csv"
        with open(caminho_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Nome", "Empresa", "Cidade"])
            for nome, empresa, cidade in sorted(faltantes):
                writer.writerow([nome, empresa, cidade])

def extrair_nomes_comprovantes_todos(pasta):
    nomes = []
    if not os.path.isdir(pasta):
        print(f"⚠️ Pasta de comprovantes não encontrada: {pasta}")
        return nomes
    try:
        for root, dirs, files in os.walk(pasta):
            for arq in files:
                if not arq.lower().endswith(".pdf"):
                    continue
                caminho = os.path.join(root, arq)
                try:
                    leitor = PdfReader(caminho, strict=False)
                    for i, pagina in enumerate(leitor.pages):
                        texto = pagina.extract_text() or ""
                        padroes_nome = [
                            r"NOME\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)",
                            r"NOME\s+DO\s+FAVORECIDO\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)",
                            r"DADOS\s+DA\s+CONTA\s+CREDITADA[^\n]*\n?\s*NOME\s*:\s*([A-ZÀ-ÿ0-9\s\-\.]+)"
                        ]
                        m = None
                        for pat in padroes_nome:
                            m = re.search(pat, texto, re.IGNORECASE | re.MULTILINE)
                            if m:
                                break
                        mag = re.search(r"AG[ÊE]NCIA\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
                        mc = re.search(r"CONTA(?:\s+CORRENTE)?\s*:\s*([\d\.\-\s]+)", texto, re.IGNORECASE)
                        pagador = extrair_info_pagador(texto)
                        ag = canonicalizar_numero(mag.group(1)) if mag else None
                        conta = canonicalizar_numero(mc.group(1)) if mc else None
                        if m:
                            nome = normalizar(m.group(1))
                            nomes.append((nome, caminho, i, ag, conta, pagador))
                except Exception as e:
                    print(f"⚠️ Erro lendo {arq}: {e}")
    except Exception as e:
        print(f"⚠️ Erro acessando {pasta}: {e}")
    return nomes

# ==========================
# 3️⃣ EXTRAIR NOMES
# ==========================
print("📄 Lendo contracheques por cidade...")
contracheques_por_cidade = extrair_nomes_contracheques(pasta_contracheques)
total_paginas = sum(len(v) for v in contracheques_por_cidade.values())
print(f"✅ {total_paginas} contracheques lidos em {len(contracheques_por_cidade)} cidades.\n")

print("📄 Lendo comprovantes...")
comprovantes = extrair_nomes_comprovantes_todos(pasta_comprovantes)
print(f"✅ {len(comprovantes)} comprovantes lidos.\n")
print("🏷️ Identificando nome da empresa a partir dos nomes dos arquivos...")
empresa_por_cidade = {}
todas_empresas = []
INVALID_TOKENS = {"RECIBO DE PAGAMENTO", "RECIBO DE PAGAMENTOS"}
for cidade, contracheques in contracheques_por_cidade.items():
    cont = Counter()
    caminhos_unicos = {c for _, c, *rest in contracheques}
    for caminho in caminhos_unicos:
        base = os.path.splitext(os.path.basename(caminho))[0]
        m = re.search(r"(?i)recibo\s+de\s+pagamento(?:s)?\s*-\s*([^-]+?)\s*-\s*([^-]+)", base)
        empresa_raw, cidade_raw = None, None
        if m:
            empresa_raw = m.group(1).strip()
            cidade_raw = m.group(2).strip()
        else:
            partes = [p.strip() for p in re.split(r"\s*-\s*", base)]
            if len(partes) >= 3:
                empresa_raw = partes[-2]
                cidade_raw = partes[-1]
            elif len(partes) == 2:
                empresa_raw = partes[0]
                cidade_raw = partes[1]
            else:
                empresa_raw = partes[0]
        empresa_norm = normalizar(empresa_raw) if empresa_raw else ""
        cidade_norm = normalizar(cidade_raw) if cidade_raw else ""
        if empresa_norm and empresa_norm not in INVALID_TOKENS:
            chave_cidade = re.sub(r'\s+', '', normalizar(cidade))
            cidade_compact = re.sub(r'\s+', '', cidade_norm)
            if not cidade_norm or cidade_compact == chave_cidade:
                cont[empresa_norm] += 1
                todas_empresas.append(empresa_norm)
    empresa_por_cidade[cidade] = cont.most_common(1)[0][0] if cont else "EMPRESA"

nome_empresa_global = Counter([e for e in todas_empresas if e not in INVALID_TOKENS]).most_common(1)[0][0] if todas_empresas else "EMPRESA"
print(f"✅ Empresa detectada: {nome_empresa_global}\n")

comp_dict_name = {nome: (pdf, pg) for nome, pdf, pg, ag, conta, _ in comprovantes}
comp_dict_acc = {}
for nome, pdf, pg, ag, conta, pagador in comprovantes:
    if ag and conta:
        comp_dict_acc[(ag, conta)] = (pdf, pg)
comp_dict_conta = {}
for nome, pdf, pg, ag, conta, pagador in comprovantes:
    if conta:
        comp_dict_conta.setdefault(conta, []).append((pdf, pg, nome, ag, conta, pagador))
index_nome2 = defaultdict(list)
for nome, pdf, pg, ag, conta, pagador in comprovantes:
    key = chave_nome2(nome)
    if key:
        index_nome2[key].append((pdf, pg, nome, ag, conta, pagador))
comp_nomes = list(comp_dict_name.keys())

# Indexação para busca rápida
word_to_names = defaultdict(set)
IGNORADOS_INDEX = {'DE', 'DA', 'DO', 'DOS', 'DAS', 'E'}
for nome in comp_nomes:
    tokens = [t for t in nome.split() if t not in IGNORADOS_INDEX]
    for token in tokens:
        word_to_names[token].add(nome)

# ==========================
# 4️⃣ GERAR PDF POR EMPRESA E GRUPO (PASTA)
# ==========================
faltando_global = []
faltantes_export = []  # NOVO: coleta para planilha
grupos = defaultdict(list)
for grupo, contracheques in contracheques_por_cidade.items():
    for nome_contra, pdf_contra, pg_contra, ag_contra, conta_contra, cidade_real in contracheques:
        base = os.path.splitext(os.path.basename(pdf_contra))[0]
        
        # LÓGICA DE NOME DA EMPRESA
        # 1. Se o grupo for uma pasta física, o nome da empresa É o nome da pasta.
        if os.path.isdir(os.path.join(pasta_contracheques, grupo)):
            emp = grupo
        else:
            # 2. Se for arquivo da raiz (grupo = cidade), tenta detectar.
            cand1 = extrair_empresa_de_nome(base, cidade_real)
            cand2 = empresa_por_cidade.get(grupo)
            cand3 = nome_empresa_global
            emp = next((c for c in (cand1, cand2, cand3) if c and c not in INVALID_TOKENS and c != "EMPRESA"), None)
            if not emp:
                emp = cand1 or "CONTRATO"

        grupos[(emp, grupo)].append((nome_contra, pdf_contra, pg_contra, ag_contra, conta_contra, cidade_real))

for (empresa_nome, grupo_nome), lista_arquivos in grupos.items():
    print(f"📂 Processando pasta: {grupo_nome} | Empresa: {empresa_nome}")
    
    # Separar Capital vs Postos vs Interior
    itens_capital = []
    itens_postos = []
    itens_interior = []
    
    # Agrupamento por município individual
    itens_por_municipio = defaultdict(list)

    # Função interna para processar e salvar uma lista de itens
    def processar_e_salvar(itens, categoria):
        if not itens:
            return
            
        paginas_local = 0
        faltando_local = []
        partes = []
        
        # Determinar nome do arquivo
        # Se a categoria for INTERIOR ou POSTOS, e o grupo for exatamente igual à cidade (arquivos da raiz),
        # usamos o nome do grupo/cidade em vez do nome genérico da categoria.
        cidades_no_grupo = {item[5] for item in itens}
        if len(cidades_no_grupo) == 1 and list(cidades_no_grupo)[0] == grupo_nome:
            sufixo_arquivo = grupo_nome
        else:
            sufixo_arquivo = categoria

        if categoria == "GOIANIA CAPITAL":
             # Mantém o padrão explícito para capital
             sufixo_arquivo = "GOIANIA CAPITAL"

        for nome_contra, pdf_contra, pg_contra, ag_contra, conta_contra, _ in itens:
            melhor_cand = None
            melhor_score = 0.0

            partes.append((pdf_contra, pg_contra))
            paginas_local += 1
            
            empresa_contra_norm = normalizar(empresa_nome) if empresa_nome else ""
            empresa_key = normalizar_contrato_chave(empresa_nome)
            base_arquivo_norm = normalizar(os.path.splitext(os.path.basename(pdf_contra))[0])


            # Primeiro: procurar por chave de 1º+2º nome
            key2 = chave_nome2(nome_contra)
            if not melhor_cand and key2 and key2 in index_nome2:
                for (cp_pdf, cp_pg, cp_nome, cp_ag, cp_conta, cp_pagador) in index_nome2[key2]:
                    is_decimo = eh_decimo_terceiro(cp_pagador)
                    s_nome = calcular_score_nomes(nome_contra, cp_nome)
                    if s_nome < FUZZY_CUTOFF:
                        continue
                    score = 100.0 * s_nome
                    if conta_contra and cp_conta and conta_contra == cp_conta:
                        score += 25.0
                    if ag_contra and cp_ag and ag_contra == cp_ag:
                        score += 15.0
                    if cp_pagador:
                        if is_decimo:
                            score += 5.0
                        elif pagador_compativel_arquivo(cp_pagador, base_arquivo_norm):
                            score += 10.0
                    if score > melhor_score:
                        melhor_score = score
                        melhor_cand = (cp_pdf, cp_pg)

            # Fallback: avaliar todos os comprovantes
            for (cp_nome, cp_pdf, cp_pg, cp_ag, cp_conta, cp_pagador) in comprovantes:
                is_decimo = eh_decimo_terceiro(cp_pagador)
                if not nome_compativel_estrito(nome_contra, cp_nome):
                    continue
                if cp_pagador and (not is_decimo) and not pagador_compativel_arquivo(cp_pagador, base_arquivo_norm):
                    continue
                s_nome = calcular_score_nomes(nome_contra, cp_nome)
                if s_nome < FUZZY_CUTOFF:
                    continue
                score = 100.0 * s_nome
                if conta_contra and cp_conta and conta_contra == cp_conta:
                    score += 25.0
                if ag_contra and cp_ag and ag_contra == cp_ag:
                    score += 15.0
                if cp_pagador:
                    if is_decimo:
                        score += 5.0
                    elif pagador_compativel_arquivo(cp_pagador, base_arquivo_norm):
                        score += 10.0
                if score > melhor_score:
                    melhor_score = score
                    melhor_cand = (cp_pdf, cp_pg)

            if melhor_cand:
                try:
                    partes.append((melhor_cand[0], melhor_cand[1]))
                    paginas_local += 1
                except Exception as e:
                    print(f"⚠️ Erro combinando {nome_contra}: {e}")
            else:
                faltando_local.append(nome_contra)
                paginas_local += 1
        
        # Salvar
        nome_saida = f"CC + CP - {empresa_nome} - {sufixo_arquivo}.pdf"
        pasta_destino = os.path.join(pasta_saida, grupo_nome)
        os.makedirs(pasta_destino, exist_ok=True)
        caminho_saida = os.path.join(pasta_destino, nome_saida)
        
        merger = PdfMerger()
        for pdf_path, pg in partes:
            try:
                merger.append(pdf_path, pages=(pg, pg+1))
            except Exception as e:
                print(f"   ⚠️ Erro anexando página {pg} de {os.path.basename(pdf_path)}: {e}")
        with open(caminho_saida, "wb") as f:
            merger.write(f)
        merger.close()
        print(f"   ✅ Salvo: {nome_saida} ({paginas_local} págs)")
        
        if faltando_local:
            print(f"   🚫 Sem comprovante em {sufixo_arquivo}:")
            for n in faltando_local:
                print(f"      - {n}")
            faltando_global.extend([(grupo_nome, n) for n in faltando_local])
            # NOVO: adicionar também à planilha de faltantes
            for n in faltando_local:
                faltantes_export.append((n, empresa_nome, sufixo_arquivo))

    for item in lista_arquivos:
        # item = (nome, pdf, pg, ag, conta, cidade_real)
        cidade_real = item[5]
        
        itens_por_municipio[cidade_real].append(item)

        if cidade_real.strip().upper() == "GOIANIA":
            itens_capital.append(item)
        elif "POSTO" in cidade_real.strip().upper():
            itens_postos.append(item)
        else:
            itens_interior.append(item)

    # 1. Gerar arquivos individuais por município
    for municipio, itens in itens_por_municipio.items():
        processar_e_salvar(itens, municipio)

    # 2. Gerar agregados (Capital / Interior) com condicional
    # "caso n tenho goiania ou que n tenha municipio e so o nome da empreso exemplo ABIN, HEMOCENTRO"
    # "nesses casos n precisa do interior ou capital"
    tem_goiania = "GOIANIA" in itens_por_municipio
    empresas_simples = {"ABIN", "HEMOCENTRO"} # Adicione outras se necessário
    eh_empresa_simples = empresa_nome in empresas_simples
    
    precisa_agregados = tem_goiania and not eh_empresa_simples

# NOVO: exportar planilha consolidada de faltantes
relatorio_xlsx = os.path.join(pasta_saida, "Relatorio_Faltantes.xlsx")
exportar_faltantes_planilha(faltantes_export, relatorio_xlsx)
print(f"📊 Planilha de faltantes gerada em: {relatorio_xlsx}")
